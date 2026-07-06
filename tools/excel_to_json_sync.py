# NANOCO Sync Tool - Excel to JSON
# Usage: python excel_to_json_sync.py "Quản lý thiết bị.xlsx" "../data"
import sys, os, re, json
from datetime import datetime, date
try:
    import openpyxl
except ImportError:
    print('Please install openpyxl: pip install openpyxl')
    raise

def clean(v):
    if v is None: return ''
    if isinstance(v, datetime): return v.strftime('%Y-%m-%d')
    if isinstance(v, date): return v.strftime('%Y-%m-%d')
    return re.sub(r'\s+', ' ', str(v).strip())

def slug(s):
    s=clean(s).lower(); s=re.sub(r'[^a-z0-9]+','-',s); return s.strip('-')[:60] or 'item'

def convert(xlsx_path, out_dir):
    wb=openpyxl.load_workbook(xlsx_path, data_only=True)
    os.makedirs(out_dir, exist_ok=True)
    ws=wb['Equipment_Register']; headers=[clean(c.value) for c in ws[1]]; equipment=[]
    for ridx,row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        rec={headers[i]: clean(row[i]) if i<len(row) else '' for i in range(len(headers))}
        name=rec.get('TÊN Name','') or rec.get('Tên Theo mã tài sản','')
        if not (name or rec.get('Mã tài sản') or rec.get('MODEL') or rec.get('Serial')): continue
        asset=rec.get('Mã tài sản',''); eid=asset or f"EQ-{ridx:04d}-{slug(name or rec.get('MODEL') or 'equipment')}"
        item={'id':eid,'assetId':asset,'category':rec.get('Phân loại',''),'assetName':rec.get('Tên Theo mã tài sản',''),'name':name,'model':rec.get('MODEL',''),'serial':rec.get('Serial',''),'parameter':rec.get('THÔNG SỐ KỸ THUẬT Parameter',''),'quantity':rec.get('SỐ LƯỢNG Quantity',''),'manufacturer':rec.get('HÃNG SX Manufacturer',''),'vendor':rec.get('NHÀ CUNG CẤP Vendor',''),'year':rec.get('Năm sản xuất',''),'area':rec.get('VỊ TRÍ LẮP ĐẶT Position',''),'maintenanceCycle':rec.get('CHU KỲ BẢO DƯỠNG Maintenance cycle',''),'installationDate':rec.get('NGÀY LẮP ĐẶT Installation date',''),'inspection':rec.get('KIỂM ĐỊNH Accreditation',''),'calibration':rec.get('Hiệu chuẩn Calibration',''),'note':rec.get('GHI CHÚ Note',''),'raw':rec}
        item['searchText']=' '.join(str(v) for v in item.values() if isinstance(v,str)).lower(); equipment.append(item)
    ws=wb['Asset_Event_Log1']; headers=[clean(c.value) for c in ws[1]]; history=[]
    for ridx,row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        rec={headers[i]: clean(row[i]) if i<len(row) else '' for i in range(len(headers))}
        if not any(rec.values()): continue
        ev={'id':rec.get('ID','') or f'EV-{ridx:04d}','startTime':rec.get('Start time',''),'completionTime':rec.get('Completion time',''),'date':rec.get('Ngày nhận thông tin','') or rec.get('Start time',''),'reporter':rec.get('Người báo thông tin cho bạn',''),'area':rec.get('Vị trí khu vực',''),'assetId':rec.get('Mã tài sản',''),'assetName':rec.get('Tên theo mã tài sản',''),'name':rec.get('Tên thiết bị',''),'model':rec.get('Model',''),'serial':rec.get('Số Serial',''),'eventType':rec.get('Loại trạng thái/sự kiện','') or 'Other','description':rec.get('Nôi dung thông tin (Giờ nhận thông tin, Kênh nhận thông báo, Mô tả lỗi hoặc tình huống)',''),'owner':rec.get('Người phụ trách thực hiện',''),'result':rec.get('Kết quả','') or rec.get('Kết quả2',''),'cost':rec.get('CHi phí',''),'downtime':rec.get('Downtime (Phút)',''),'rootCause':rec.get('Nguyên Nhân',''),'note':rec.get('Ghi chú',''),'raw':rec}
        ev['searchText']=' '.join(str(v) for v in ev.values() if isinstance(v,str)).lower(); history.append(ev)
    meta={'sourceFile':os.path.basename(xlsx_path),'generatedAt':datetime.now().isoformat(timespec='seconds'),'equipmentCount':len(equipment),'historyCount':len(history),'sheets':['Equipment_Register','Asset_Event_Log1']}
    for fn,obj in [('equipment.json',equipment),('history.json',history),('meta.json',meta)]:
        with open(os.path.join(out_dir,fn),'w',encoding='utf-8') as f: json.dump(obj,f,ensure_ascii=False,indent=2)
    print('SYNC DONE:', meta)
if __name__=='__main__':
    xlsx=sys.argv[1] if len(sys.argv)>1 else 'Quản lý thiết bị.xlsx'
    out=sys.argv[2] if len(sys.argv)>2 else 'data'
    convert(xlsx,out)
