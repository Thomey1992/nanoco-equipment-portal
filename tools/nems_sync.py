# NEMS Sync Tool - Nguyễn Lĩnh (Thomey)
# Reads the finalized Excel workbook and generates data/equipment.json + data/history.json.
# Optional: upload those JSON files directly to GitHub using a GitHub token.

import base64
import datetime as dt
import getpass
import json
import os
import re
import sys
import urllib.request
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

REPO = "Thomey1992/nanoco-equipment-portal"
BRANCH = "main"
ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"


def clean(v):
    if v is None:
        return ""
    if isinstance(v, dt.datetime):
        return v.date().isoformat()
    if isinstance(v, dt.date):
        return v.isoformat()
    return re.sub(r"\s+", " ", str(v).strip())


def norm(s):
    return re.sub(r"\s+", " ", str(s or "").replace("\n", " ").strip())


def load_workbook(path):
    return openpyxl.load_workbook(path, data_only=True)


def parse_equipment(wb):
    ws = wb["Equipment_Register"]
    headers = [norm(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        row = {headers[i]: vals[i] for i in range(len(headers))}
        no = clean(row.get("No."))
        name = clean(row.get("TÊN Name"))
        if not no or not re.match(r"^\d+(\.0)?$", no) or not name:
            continue
        asset = clean(row.get("Mã tài sản"))
        eid = asset or f"EQ-{int(float(no)):04d}"
        rows.append({
            "id": eid,
            "no": no,
            "category": clean(row.get("Phân loại")),
            "assetCode": asset,
            "assetName": clean(row.get("Tên Theo mã tài sản")),
            "name": name,
            "model": clean(row.get("MODEL")),
            "serial": clean(row.get("Serial")),
            "parameter": clean(row.get("THÔNG SỐ KỸ THUẬT Parameter")),
            "quantity": clean(row.get("SỐ LƯỢNG Quantity")),
            "manufacturer": clean(row.get("HÃNG SX Manufacturer")),
            "vendor": clean(row.get("NHÀ CUNG CẤP Vendor")),
            "year": clean(row.get("Năm sản xuất")),
            "position": clean(row.get("VỊ TRÍ LẮP ĐẶT Position")),
            "area": clean(row.get("VỊ TRÍ LẮP ĐẶT Position")),
            "pmCycle": clean(row.get("CHU KỲ BẢO DƯỠNG Maintenance cycle")),
            "installDate": clean(row.get("NGÀY LẮP ĐẶT Installation date")),
            "inspection": clean(row.get("KIỂM ĐỊNH Accreditation")),
            "calibration": clean(row.get("Hiệu chuẩn Calibration")),
            "note": clean(row.get("GHI CHÚ Note")),
            "sparePart": clean(row.get("Phụ tùng")),
            "maintenanceLink": clean(row.get("Link hướng dẫn bảo trì")),
            "documentLink": clean(row.get("Link tài liệu hồ sơ")),
            "operationLink": clean(row.get("Link hướng dẫn vận hành")),
            "calibrationLink": clean(row.get("Link hiệu chuẩn")),
            "inspectionLink": clean(row.get("Link kiểm định")),
            "status": "Running"
        })
    return rows


def parse_history(wb, equipment):
    ws = wb["Asset_Event_Log1"]
    headers = [norm(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        row = {headers[i]: vals[i] for i in range(len(headers))}
        event_id = clean(row.get("ID"))
        if not event_id:
            continue
        asset = clean(row.get("Mã tài sản"))
        name = clean(row.get("Tên thiết bị")) or clean(row.get("Tên theo mã tài sản"))
        model = clean(row.get("Model"))
        eqid = asset
        if not eqid:
            lname = name.lower()
            for e in equipment:
                if lname and (lname in e["name"].lower() or e["name"].lower() in lname):
                    eqid = e["id"]
                    break
            if not eqid and model:
                for e in equipment:
                    if model.lower() == e["model"].lower():
                        eqid = e["id"]
                        break
        if not eqid:
            eqid = f"EVENT-{event_id}"
        rows.append({
            "eventId": event_id,
            "assetId": eqid,
            "date": clean(row.get("Ngày nhận thông tin")) or clean(row.get("Start time")),
            "startTime": clean(row.get("Start time")),
            "completionTime": clean(row.get("Completion time")),
            "email": clean(row.get("Email")),
            "name": clean(row.get("Name")),
            "reporter": clean(row.get("Người báo thông tin cho bạn")),
            "area": clean(row.get("Vị trí khu vực")),
            "machineName": name,
            "model": model,
            "serial": clean(row.get("Số Serial")),
            "type": clean(row.get("Loại trạng thái/sự kiện")) or "Other",
            "description": clean(row.get("Nôi dung thông tin (Giờ nhận thông tin, Kênh nhận thông báo, Mô tả lỗi hoặc tình huống)")),
            "attachment": clean(row.get("Hình ảnh, video hiện trạng thiết bị, sự cố nếu có, hoặc ảnh màn hình tin nhắn cuộc gọi báo sự cố")),
            "owner": clean(row.get("Người phụ trách thực hiện")),
            "result": clean(row.get("Kết quả")) or clean(row.get("Kết quả2")),
            "cost": clean(row.get("CHi phí")),
            "downtime": clean(row.get("Downtime (Phút)")),
            "cause": clean(row.get("Nguyên Nhân")),
            "note": clean(row.get("Ghi chú"))
        })
    return rows


def infer_status(equipment, history):
    latest = {}
    for h in history:
        latest[h["assetId"]] = h
    for e in equipment:
        h = latest.get(e["id"]) or latest.get(e.get("assetCode", ""))
        if not h:
            continue
        txt = " ".join([h.get("type", ""), h.get("description", ""), h.get("result", "")]).lower()
        if re.search(r"thanh lý|scrap|disposal", txt):
            e["status"] = "Waiting Disposal"
        elif re.search(r"chờ.*sửa|waiting repair", txt):
            e["status"] = "Waiting Repair"
        elif re.search(r"hư|dừng|breakdown|repair|sửa", txt):
            e["status"] = "Breakdown"


def write_json(equipment, history):
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    (DATA_DIR / "equipment.json").write_text(json.dumps(equipment, ensure_ascii=False, indent=2), encoding="utf-8")
    (DATA_DIR / "history.json").write_text(json.dumps(history, ensure_ascii=False, indent=2), encoding="utf-8")


def github_get_sha(path, token):
    url = f"https://api.github.com/repos/{REPO}/contents/{path}?ref={BRANCH}"
    req = urllib.request.Request(url, headers={"Authorization": f"Bearer {token}", "Accept": "application/vnd.github+json"})
    try:
        with urllib.request.urlopen(req) as r:
            return json.loads(r.read().decode()).get("sha")
    except Exception:
        return None


def github_put(path, local_file, token):
    content = base64.b64encode(Path(local_file).read_bytes()).decode()
    sha = github_get_sha(path, token)
    body = {"message": f"NEMS sync {path}", "content": content, "branch": BRANCH}
    if sha:
        body["sha"] = sha
    req = urllib.request.Request(
        f"https://api.github.com/repos/{REPO}/contents/{path}",
        data=json.dumps(body).encode(),
        method="PUT",
        headers={"Authorization": f"Bearer {token}", "Accept": "application/vnd.github+json", "Content-Type": "application/json"}
    )
    with urllib.request.urlopen(req) as r:
        return r.status


def main():
    print("=== NANOCO Equipment Sync Tool ===")
    excel_path = input("Kéo-thả file Excel vào đây rồi Enter: ").strip().strip('"')
    if not excel_path or not Path(excel_path).exists():
        print("Không tìm thấy file Excel.")
        input("Enter để thoát...")
        return
    wb = load_workbook(excel_path)
    equipment = parse_equipment(wb)
    history = parse_history(wb, equipment)
    infer_status(equipment, history)
    write_json(equipment, history)
    print(f"Đã tạo JSON: {len(equipment)} thiết bị, {len(history)} sự kiện.")
    print(f"File: {DATA_DIR / 'equipment.json'}")
    print(f"File: {DATA_DIR / 'history.json'}")
    ans = input("Bạn muốn tự upload JSON lên GitHub không? (y/n): ").strip().lower()
    if ans == "y":
        token = os.environ.get("GITHUB_TOKEN") or getpass.getpass("Nhập GitHub token: ")
        github_put("data/equipment.json", DATA_DIR / "equipment.json", token)
        github_put("data/history.json", DATA_DIR / "history.json", token)
        print("Đã upload lên GitHub. Đợi GitHub Pages deploy 30-90 giây, rồi bấm Refresh trên web.")
    else:
        print("Bạn có thể upload thủ công 2 file trong thư mục data lên GitHub.")
    input("Hoàn tất. Enter để thoát...")


if __name__ == "__main__":
    main()
